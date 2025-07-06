import openpyxl
import os

# Script konumunu al
base_dir = os.path.dirname(os.path.abspath(__file__))

# Excel dosyasının yolunu oluştur
dosya_yolu = os.path.join(base_dir, "veriler.xlsx")

# Dosya var mı kontrol et
if not os.path.exists(dosya_yolu):
    print(f"{dosya_yolu} dosyası bulunamadı.")
    exit()

# Excel dosyasını aç
wb = openpyxl.load_workbook(dosya_yolu)
ws = wb.active  # Aktif çalışma sayfasını seç

# İlk satırdaki başlıkları al
basliklar = [cell.value for cell in ws[1]]
veriler = []

# 2. satırdan başlayarak tüm veriyi oku
for row in ws.iter_rows(min_row=2, values_only=True):
    veriler.append(list(row))

# Verileri ilk sütuna (ID) göre sırala
sirali_veriler = sorted(veriler, key=lambda x: int(x[0]))

# Yeni Excel dosyası oluştur
yeni_wb = openpyxl.Workbook()
yeni_ws = yeni_wb.active

# Başlıkları yeni dosyaya ekle
yeni_ws.append(basliklar)

# Sıralanmış verileri yeni dosyaya ekle
for satir in sirali_veriler:
    yeni_ws.append(satir)

# Yeni dosyayı kaydet
yeni_wb.save(os.path.join(base_dir, "sirali_veriler.xlsx"))

print("Veriler başarıyla sıralandı ve yeni dosyaya kaydedildi: sirali_veriler.xlsx")