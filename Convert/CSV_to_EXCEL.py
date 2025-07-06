import csv
import openpyxl
import os
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Script konumunu al
base_dir = os.path.dirname(os.path.abspath(__file__))

# Dosya yollarını oluştur
csv_dosyasi = os.path.join(base_dir, "employees.csv")
excel_dosyasi = os.path.join(base_dir, "veriler.xlsx")

# CSV dosyası var mı kontrol et
if not os.path.exists(csv_dosyasi):
    print(f"{csv_dosyasi} dosyası bulunamadı.")
    exit()

try:
    # CSV dosyasını oku
    with open(csv_dosyasi, "r", encoding="utf-8") as f:
        reader = list(csv.reader(f))

        # Dosya boş mu kontrol et
        if not reader:
            print(f"{csv_dosyasi} dosyası boş.")
            exit()
        
        # Başlıkları al ve veriyi ID'ye göre sırala
        basliklar = reader[0]
        veriler = sorted(reader[1:], key=lambda x: int(x[0]))

    # Yeni Excel dosyası oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Çalışanlar"

    # Başlıkları ve veriyi Excel'e ekle
    ws.append(basliklar)
    for satir in veriler:
        ws.append(satir)

    # Tablo oluşturmak için veri aralığını belirle
    son_satir = ws.max_row
    son_sutun = ws.max_column
    sutun_harfleri = [openpyxl.utils.get_column_letter(i) for i in range(1, son_sutun + 1)]
    veri_araligi = f"A1:{sutun_harfleri[-1]}{son_satir}"

    # Excel tablosu oluştur ve stil uygula
    tablo = Table(displayName="ÇalışanlarTablosu", ref=veri_araligi)
    stil = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,  # Zebra çizgiler
        showColumnStripes=False
    )
    tablo.tableStyleInfo = stil
    ws.add_table(tablo)

    # Sütun genişliklerini otomatik ayarla
    for col_num in range(1, son_sutun + 1):
        column = ws.column_dimensions[get_column_letter(col_num)]
        max_uzunluk = 0

        # Her sütundaki en uzun değeri bul
        for row_num in range(1, min(son_satir + 1, 1000)):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                max_uzunluk = max(max_uzunluk, len(str(cell.value)))

        # Sütun genişliğini ayarla (minimum 5, maksimum 50)
        column.width = min(max_uzunluk + 5, 50)

    # Excel dosyasını kaydet
    wb.save(excel_dosyasi)
    print(f"CSV dosyası {csv_dosyasi} başarıyla {excel_dosyasi} olarak kaydedildi.")

except Exception as e:
    print(f"Bir hata oluştu: {e}")
    exit()