import os
import csv
import openpyxl

# Script konumunu al
base_dir = os.path.dirname(os.path.abspath(__file__))

# Dosya yollarını oluştur
csv_dosyasi = os.path.join(base_dir, "employees2.csv")
excel_dosyasi = os.path.join(base_dir, "veriler.xlsx")

# Excel dosyası var mı kontrol et
if not os.path.exists(excel_dosyasi):
    print(f"{excel_dosyasi} dosyası bulunamadı.")
    exit()

# Excel dosyasını aç
wb = openpyxl.load_workbook(excel_dosyasi)
ws = wb.active  # Aktif çalışma sayfasını seç

# İlk satırdaki başlıkları al
basliklar = [cell.value for cell in ws[1]]
veriler = []

# 2. satırdan başlayarak tüm veriyi oku
for row in ws.iter_rows(min_row=2, values_only=True):
    veriler.append(list(row))

# Verileri ilk sütuna (ID) göre sırala
sirali_veriler = sorted(veriler, key=lambda x: int(x[0]))

# Sıralanmış veriyi CSV dosyasına yaz
with open(csv_dosyasi, mode='w', encoding='utf-8', newline='') as csv_file:
    writer = csv.writer(csv_file)
    writer.writerow(basliklar)  # Başlıkları yaz
    writer.writerows(sirali_veriler)  # Veriyi yaz

print(f"Veriler başarıyla {csv_dosyasi} dosyasına kaydedildi.")