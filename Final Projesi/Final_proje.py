import csv, os, sys, logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Union
from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Gerekli kütüphaneleri kontrol et ve yükle
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Hata: openpyxl kütüphanesi gerekli. Şu komutla yükleyin: pip install openpyxl")
    sys.exit(1)


class CSVToExcelConverter:
    """
    CSV'den Excel'e dönüştürücü sınıfı - Profesyonel sürüm
    
    Bu sınıf, birden fazla CSV/TXT dosyasını Excel formatına dönüştürme 
    işlevselliği sağlar. Tek veya çoklu çalışma sayfası seçenekleri sunar.
    """
    
    # Sınıf sabitleri - değiştirilebilir ayarlar
    SUPPORTED_EXTENSIONS = ('.csv', '.txt') # Desteklenen dosya uzantıları
    OUTPUT_DIR = 'output'  # Çıktı dosyalarının kaydedileceği klasör
    LOG_FORMAT = '%(asctime)s - %(levelname)s - %(message)s' # Log formatı
    MAX_EXCEL_SHEET_NAME_LENGTH = 31  # Excel'in maksimum sayfa adı uzunluğu
    
    def __init__(self, output_directory: str = OUTPUT_DIR, log_level: str = 'INFO'):
        """
        Dönüştürücüyü yapılandırmayla başlat
        
        Args:
            output_directory (str): Çıktı dosyaları için klasör
            log_level (str): Log seviyesi (DEBUG, INFO, WARNING, ERROR)
        """
        # Temel dizin ayarları - scriptin bulunduğu konumu al
        self.base_dir = Path(os.path.dirname(os.path.abspath(__file__)))
        self.output_dir = self.base_dir / output_directory
        self.current_dir = self.base_dir
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Dizinleri ve loglama sistemini kurulum
        self._setup_logging(log_level)
        self._setup_directories()
        
        self.logger.info(f"Çıktı dizini: {self.output_dir.absolute()}")
    
    def _setup_directories(self) -> None:
        """Gerekli dizinleri yoksa oluştur"""
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            if hasattr(self, 'logger'):
                self.logger.debug(f"Çıktı dizini oluşturuldu: {self.output_dir}")
        except Exception as e:
            raise RuntimeError(f"Çıktı dizini oluşturulamadı: {e}")
    
    def _setup_logging(self, log_level: str) -> None:
        """
        Uygulama için loglama sistemini yapılandır
        
        Args:
            log_level (str): Log seviyesi
        """
        temp_logger = logging.getLogger(__name__)

        # Önce çıktı dizinini oluşturmaya çalış
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(f"Uyarı: Çıktı dizini oluşturulamadı: {e}")
            self.output_dir = self.base_dir

        # Log dosyası için tam yol
        log_file = self.output_dir / f'conversion_{self.timestamp}.log'
        
        # Loglama sistemini yapılandır - hem dosyaya hem konsola yaz
        logging.basicConfig(
            level=getattr(logging, log_level.upper()),
            format=self.LOG_FORMAT,
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Loglama başlatıldı - Seviye: {log_level}")
    
    def discover_csv_files(self, directory: Optional[str] = None) -> List[Path]:
        """
        Belirtilen dizinde veya ana dizinde CSV ve TXT dosyalarını keşfet

        Args:
            directory (Optional[str]): Aranacak dizin (base_dir'e göre)
    
        Returns:
            List[Path]: Bulunan CSV/TXT dosyalarının listesi
        """
        try:
            if directory:
                search_dir = self.base_dir / directory
                if not search_dir.exists():
                    search_dir = Path(directory) # Mutlak yol olarak dene
            else:
                search_dir = self.base_dir

            # Dizin var mı kontrol et
            if not search_dir.exists():
                self.logger.error(f"Dizin bulunamadı: {search_dir}")
                return []
            
            # Desteklenen dosya uzantılarını ara
            files = []
            for file_path in search_dir.iterdir():
                if (file_path.is_file() and 
                    file_path.suffix.lower() in self.SUPPORTED_EXTENSIONS):
                    files.append(file_path)
            
            self.logger.info(f"{search_dir} dizininde {len(files)} CSV/TXT dosyası bulundu")
            return sorted(files) # Alfabetik sırala
            
        except Exception as e:
            self.logger.error(f"Dosya keşif hatası: {e}")
            return []
        
    def get_directory_selection(self) -> Optional[str]:
        """
        Kullanıcıdan dizin seçimi al
    
        Returns:
            Optional[str]: Seçilen dizin yolu veya mevcut dizin için None
        """
        print(f"\n{'='*60}")
        print("DİZİN SEÇİMİ")
        print(f"{'='*60}")

        print(f"Ana Dizin: {self.base_dir}")
    
        # Mevcut alt dizinleri listele (output hariç)
        subdirs = [d for d in self.base_dir.iterdir() if d.is_dir() and d.name != self.OUTPUT_DIR]

        if subdirs:
            print("Mevcut alt dizinler:")
            print("0. Ana Dizin (script konumu)")
            for idx, dir_path in enumerate(subdirs, 1):
                print(f"{idx}. {dir_path.name}")
            
            print("\nSeçenekler:")
            print("• Mevcut dizin: ENTER tuşuna basın veya 0 yazın")
            print("• Numara ile seçim: 1, 2, 3...")
            print("• Özel yol: belgeler2, docs, vb.")

            user_input = input("\n Dizin seçin: ").strip()

            # Boş veya 0 ise mevcut dizin
            if not user_input or user_input == '0':
                return None
            
            # Numara ile seçim
            if user_input.isdigit():
                idx = int(user_input) - 1
                if 0 <= idx < len(subdirs):
                    return subdirs[idx].name
            else:
                # Özel path girişi
                return user_input
            
        else:
            print("Mevcut dizinde alt dizin bulunamadı.")
            custom_path = input("Özel yol girin (veya mevcut dizin için ENTER): ").strip()
            return custom_path if custom_path else None
    
    def get_user_file_selection(self, available_files: List[Path]) -> List[Path]:
        """
        Kullanıcıdan dosya seçimi al ve doğrula
        
        Args:
            available_files (List[Path]): Mevcut dosyaların listesi
            
        Returns:
            List[Path]: Seçilen dosyalar
        """
        if not available_files:
            print("\nMevcut dizinde CSV/TXT dosyası bulunamadı")
            return []
        
        print(f"\n{'='*60}")
        print("📁 MEVCUT CSV/TXT DOSYALARI")
        print(f"{'='*60}")
        
        # Dosyaları boyutlarıyla birlikte listele
        for idx, file_path in enumerate(available_files, 1):
            file_size = self._format_file_size(file_path.stat().st_size)
            print(f"{idx:2d}. {file_path.name:<35} ({file_size})")
        
        print(f"\n{'Seçim Seçenekleri:':<20}")
        print("  • Tüm dosyalar: ENTER tuşuna basın")
        print("  • Numaralarla: 1,3,5")
        print("  • İsimlerle: dosya1.csv,dosya2.txt")
        
        user_input = input("\nSeçiminiz: ").strip()
        
        # Boş ise tüm dosyaları seç
        if not user_input:
            self.logger.info("Kullanıcı tüm dosyaları seçti")
            return available_files
        
        # Kullanıcı girişini çözümle
        selected_files = self._parse_user_selection(user_input, available_files)
        self.logger.info(f"Kullanıcı {len(selected_files)} dosya seçti: {[f.name for f in selected_files]}")
        
        return selected_files
    
    def _parse_user_selection(self, user_input: str, available_files: List[Path]) -> List[Path]:
        """
        Kullanıcı seçim girişini çözümle
        
        Args:
            user_input (str): Kullanıcı giriş metni
            available_files (List[Path]): Mevcut dosyalar
            
        Returns:
            List[Path]: Çözümlenen seçilen dosyalar
        """
        selected_files = []
        
        # Virgülle ayrılmış seçimleri işle
        for selection in user_input.split(','):
            selection = selection.strip()
            
            if selection.isdigit():
                # Numara ile seçim
                idx = int(selection) - 1
                if 0 <= idx < len(available_files):
                    selected_files.append(available_files[idx])
                else:
                    self.logger.warning(f"Geçersiz dosya numarası: {selection}")
            else:
                # Dosya adı ile seçim
                matching_files = [f for f in available_files if f.name == selection]
                if matching_files:
                    selected_files.extend(matching_files)
                else:
                    self.logger.warning(f"Dosya bulunamadı: {selection}")
        
        # Tekrarları kaldır ama sırayı koru
        return list(dict.fromkeys(selected_files))
    
    def get_excel_format_preference(self) -> str:
        """
        Kullanıcıdan Excel format tercihi al
        
        Returns:
            str: Format tercihi ('single' veya 'multiple')
        """
        print(f"\n{'='*60}")
        print("EXCEL FORMAT SEÇENEKLERİ")
        print(f"{'='*60}")
        print("[1] Tek çalışma sayfası - Tüm veriler birleştirilir")
        print("[2] Çoklu çalışma sayfası - Her dosya için ayrı sayfa")
        
        while True:
            choice = input("\n📝 Format seçin (1 veya 2): ").strip()
            if choice == '1':
                self.logger.info("Kullanıcı tek çalışma sayfası formatını seçti")
                return 'single'
            elif choice == '2':
                self.logger.info("Kullanıcı çoklu çalışma sayfası formatını seçti")
                return 'multiple'
            else:
                print("Lütfen 1 veya 2 girin")

    def get_sort_preference(self) -> bool:
        """Kullanıcıdan ID sıralama tercihi al"""
        print(f"\n{'='*60}")
        print("SIRALAMA SEÇENEKLERİ")
        print(f"{'='*60}")
        print("[1] ID sütununa göre sırala (varsa)")
        print("[2] Orijinal sırayı koru")

        while True:
            choice = input("\n📝 Sıralama seçeneği seçin (1 veya 2): ").strip()
            if choice == '1':
                self.logger.info("Kullanıcı ID sıralamasını seçti")
                return True
            elif choice == '2':
                self.logger.info("Kullanıcı orijinal sırayı seçti")
                return False
            else:
                print("Lütfen 1 veya 2 girin")
    
    def read_csv_file(self, file_path: Path, sorted_by_id: bool = False) -> Optional[List[List[str]]]:
        """
        CSV dosyasını oku ve hata yönetimi ile çözümle
        
        Args:
            file_path (Path): CSV dosyasının yolu
            sorted_by_id (bool): ID'ye göre sıralanacak mı
            
        Returns:
            Optional[List[List[str]]]: Çözümlenen CSV verisi veya hata durumunda None
        """
        try:
            with open(file_path, 'r', encoding='utf-8', newline='') as file:
                # Ayırıcı karakteri tespit et
                sample = file.read(1024)
                file.seek(0)
                
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                # CSV okuyucu oluştur ve veriyi çek
                reader = csv.reader(file, delimiter=delimiter)
                data = list(reader)
                
                if data:
                    # ID sıralaması isteniyorsa uygula
                    if sorted_by_id:
                        sorted_data = self._sort_data_by_id(data)
                        self.logger.debug(f"{file_path.name} başarıyla okundu: {len(sorted_data)} satır")
                        return sorted_data
                    else:
                        self.logger.debug(f"{file_path.name} başarıyla okundu: {len(data)} satır")
                        return data
                else:
                    self.logger.warning(f"Boş dosya: {file_path.name}")
                    return None
                    
        except UnicodeDecodeError:
            # Farklı kodlamaları dene
            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    with open(file_path, 'r', encoding=encoding, newline='') as file:
                        reader = csv.reader(file)
                        data = list(reader)
                        self.logger.debug(f"{file_path.name} {encoding} kodlaması ile okundu")
                        return data
                except UnicodeDecodeError:
                    continue
            
            self.logger.error(f"Dosya kodlaması çözülemedi: {file_path.name}")
            return None
            
        except Exception as e:
            self.logger.error(f"{file_path.name} okuma hatası: {e}")
            return None
        
    def _sort_data_by_id(self, data: List[List[str]]) -> List[List[str]]:
        """
        CSV verisini ID sütununa göre sırala
        
        Args:
            data (List[List[str]]): Ham CSV verisi
            
        Returns:
            List[List[str]]: Sıralanmış CSV verisi
        """
        try:
            # Sadece header varsa sıralama yapma
            if len(data) <= 1:
                return data
            
            header = data[0]
            rows = data[1:]

            # ID sütununu ara
            id_column_index = None
            possible_id_names = ['ID', 'id', 'Id', 'iD', 'EmployeeID', 'employee_id', 'employeeId', 
                               'kimlik_no', 'KimlikNo', 'NO', 'no', 'No', 'kimlik', 'KIMLIK']

            for i, column_name in enumerate(header):
                if column_name.strip() in possible_id_names:
                    id_column_index = i
                    break

            # ID sütunu bulunamazsa sıralama yapma
            if id_column_index is None:
                self.logger.debug("ID sütunu bulunamadı, sıralama atlanıyor")
                return data
            
            # Sıralama anahtarı fonksiyonu
            def sort_key(row):
                try:
                    id_value = row[id_column_index].strip()
                    # Sayısal ID ise integer'a çevir
                    if id_value.isdigit():
                        return int(id_value)
                    else:
                        return id_value.lower()
                except (IndexError, ValueError):
                    return float('inf')  # Hatalı satırları sona at
                
            # Sıralama işlemi
            sorted_rows = sorted(rows, key=sort_key)
            sorted_data = [header] + sorted_rows

            self.logger.debug(f"Veri '{header[id_column_index]}' sütununa göre sıralandı (indeks {id_column_index})")
            return sorted_data

        except Exception as e:
            self.logger.error(f"ID'ye göre sıralama yapılamadı: {e}")
            return data
    
    def create_excel_workbook(self, files_data: dict, format_type: str) -> Optional[openpyxl.Workbook]:
        """
        CSV verisinden Excel çalışma kitabı oluştur
        
        Args:
            files_data (dict): dosya_adı: veri çiftlerinin sözlüğü
            format_type (str): 'single' veya 'multiple'
            
        Returns:
            Optional[openpyxl.Workbook]: Oluşturulan çalışma kitabı veya hata durumunda None
        """
        try:
            workbook = openpyxl.Workbook()
            
            # Format tipine göre çalışma sayfası oluştur
            if format_type == 'single':
                self._create_single_worksheet(workbook, files_data)
            else:
                self._create_multiple_worksheets(workbook, files_data)
            
            self.logger.info(f"Excel çalışma kitabı {len(workbook.worksheets)} sayfa ile oluşturuldu")
            return workbook
            
        except Exception as e:
            self.logger.error(f"Excel çalışma kitabı oluşturma hatası: {e}")
            return None
    
    def _create_single_worksheet(self, workbook: openpyxl.Workbook, files_data: dict) -> None:
        """Birleştirilmiş verilerle tek çalışma sayfası oluştur"""
        worksheet = workbook.active
        worksheet.title = "Combined_Data"
        
        header_written = False
        total_rows = 0
        
        # Tüm dosyaları tek sayfada birleştir
        for file_path, data in files_data.items():
            if not data:
                continue
            
            # Header'ı sadece bir kez yaz
            if not header_written and data:
                self._write_row_with_formatting(worksheet, data[0], is_header=True)
                header_written = True
                total_rows += 1
            
            # Veri satırlarını ekle
            for row in data[1:]:
                worksheet.append(row)
                total_rows += 1
            
            self.logger.debug(f"{file_path.name} dosyasından {len(data)-1} veri satırı eklendi")
        
        # Sayfa formatını uygula
        self._apply_worksheet_formatting(worksheet)
        worksheet.freeze_panes = 'A2'  # Header satırını dondur
        self.logger.info(f"Tek çalışma sayfası toplam {total_rows} satır ile oluşturuldu")
    
    def _create_multiple_worksheets(self, workbook: openpyxl.Workbook, files_data: dict) -> None:
        """Her dosya için ayrı çalışma sayfası oluştur"""
        # Varsayılan çalışma sayfasını kaldır
        workbook.remove(workbook.active)
        
        # Her dosya için ayrı sayfa oluştur
        for file_path, data in files_data.items():
            if not data:
                continue
            
            # Güvenli sayfa adı oluştur
            sheet_name = self._create_safe_sheet_name(file_path.stem)
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Tüm veriyi yaz
            for idx, row in enumerate(data):
                if idx == 0:
                    self._write_row_with_formatting(worksheet, row, is_header=True)
                else:
                    self._write_row_with_formatting(worksheet, row, is_header=False)
            
            # Sayfa formatını uygula
            self._apply_worksheet_formatting(worksheet)
            worksheet.freeze_panes = 'A2'  # Header satırını dondur
            self.logger.debug(f"'{sheet_name}' çalışma sayfası {len(data)} satır ile oluşturuldu")
    
    def _write_row_with_formatting(self, worksheet, row_data: List[str], is_header: bool = False) -> None:
        """Opsiyonel header formatıyla satır yaz"""
        worksheet.append(row_data)

        # Kenarlık stilleri
        header_border = Border(
            left=Side(style='medium', color='1F4788'),
            right=Side(style='medium', color='1F4788'),
            top=Side(style='medium', color='1F4788'),
            bottom=Side(style='medium', color='1F4788')
        )

        data_border = Border(
            left=Side(style='thin', color='E1E5E9'),
            right=Side(style='thin', color='E1E5E9'),
            top=Side(style='thin', color='E1E5E9'),
            bottom=Side(style='thin', color='E1E5E9')
        )
        
        if is_header:
            # Header formatlaması uygula
            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in worksheet[worksheet.max_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
        else:
            # Veri satırı formatlaması
            data_font = Font(name='Calibri', size=10, color='333333')
            row_num = worksheet.max_row
        
            # Zebra çizgiler için alternatif renkler
            if row_num % 2 == 0:
                data_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            else:
                data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            for cell in worksheet[worksheet.max_row]:
                cell.font = data_font
                cell.fill = data_fill
                cell.border = data_border

                # Sayısal değerleri sağa hizala
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if abs(cell.value) >= 1000:
                        cell.number_format = '#,##0.00'  # Binlik ayracı
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def _apply_worksheet_formatting(self, worksheet) -> None:
        """Çalışma sayfasına genel formatlama uygula"""
        # Sütun genişliklerini otomatik ayarla
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # En uzun içeriği bul
            for cell in column:
                try:
                    cell_length = len(str(cell.value or "")) + 2  # Görünürlük için padding ekle
                    max_length = max(max_length, cell_length) + 0.09
                except:
                    pass
            
            # Sütun genişliğini makul sınırlarla ayarla
            adjusted_width = min(max(max_length + 5, 15), 55)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Excel tablosu oluştur
        self._create_table(worksheet)

    def _create_table(self, worksheet) -> None:
        """Excel tablosu oluştur"""
        try:
            last_row = worksheet.max_row
            last_column = worksheet.max_column

            # Veri varsa tablo oluştur
            if last_row > 1 and last_column > 0:
                column_letters = [get_column_letter(i) for i in range(1, last_column + 1)]
                table_range = f"A1:{column_letters[-1]}{last_row}"

                # Tablo oluştur ve stil uygula
                table = Table(displayName= f"Table_{worksheet.title}", ref=table_range)

                style = TableStyleInfo(
                    name="TableStyleLight9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,  # Çizgili satırlar
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                worksheet.add_table(table)

                self.logger.debug(f"Excel tablosu oluşturuldu: {table_range}")
        
        except Exception as e:
            self.logger.warning(f"Excel tablosu oluşturulamadı: {e}")
    
    def _create_safe_sheet_name(self, name: str) -> str:
        """Güvenli Excel çalışma sayfası adı oluştur"""
        # Geçersiz karakterleri kaldır
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        safe_name = name
        for char in invalid_chars:
            safe_name = safe_name.replace(char, '_')
        
        # Excel'in sınırına göre kısalt
        return safe_name[:self.MAX_EXCEL_SHEET_NAME_LENGTH]
    
    def save_excel_file(self, workbook: openpyxl.Workbook) -> Optional[Path]:
        """
        Excel çalışma kitabını dosyaya kaydet
        
        Args:
            workbook (openpyxl.Workbook): Kaydedilecek çalışma kitabı
            
        Returns:
            Optional[Path]: Kaydedilen dosyanın yolu veya hata durumunda None
        """
        try:
            excel_filename = f"excel_rapor_{self.timestamp}.xlsx"
            excel_path = self.output_dir / excel_filename
            
            workbook.save(excel_path)
            self.logger.info(f"Excel dosyası kaydedildi: {excel_path}")
            
            return excel_path
            
        except Exception as e:
            self.logger.error(f"Excel dosyası kaydetme hatası: {e}")
            return None
    
    def create_archive(self, source_files: List[Path], excel_file: Path) -> Optional[Path]:
        """
        Kaynak dosyalar ve Excel çıktısı ile ZIP arşivi oluştur
        
        Args:
            source_files (List[Path]): Kaynak CSV dosyaları
            excel_file (Path): Oluşturulan Excel dosyası
            
        Returns:
            Optional[Path]: Oluşturulan arşivin yolu veya hata durumunda None
        """
        try:
            archive_filename = f"csv_arsiv_{self.timestamp}.zip"
            archive_path = self.output_dir / archive_filename
            
            # ZIP arşivi oluştur
            with ZipFile(archive_path, 'w', compression=ZIP_DEFLATED, compresslevel=6) as zipf:
                # Kaynak dosyaları ekle
                for file_path in source_files:
                    if file_path.exists():
                        zipf.write(file_path, file_path.name)
                
                # Excel dosyasını ekle
                if excel_file and excel_file.exists():
                    zipf.write(excel_file, excel_file.name)
                
                # Log dosyasını ekle
                log_files = list(self.output_dir.glob('conversion_*.log'))
                if log_files:
                    zipf.write(log_files[-1], log_files[-1].name)
            
            self.logger.info(f"Arşiv oluşturuldu: {archive_path}")
            return archive_path
            
        except Exception as e:
            self.logger.error(f"Arşiv oluşturma hatası: {e}")
            return None
    
    def _format_file_size(self, size_bytes: int) -> str:
        """Dosya boyutunu okunabilir formatta düzenle"""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.1f} TB"
    
    def run_conversion(self) -> bool:
        """
        Tam dönüştürme sürecini çalıştır
        
        Returns:
            bool: Başarılıysa True, aksi halde False
        """
        try:
            # Başlık ve versiyon bilgisi
            print(f"\n{'='*60}")
            print("CSV'DEN EXCEL'E DÖNÜŞTÜRÜCÜ - Profesyonel Sürüm")
            print(f"{'='*60}")
            print(f"Versiyon: 1.0.0 | Zaman Damgası: {self.timestamp}")

            # Adım 1: Dizin seçimi
            selected_directory = self.get_directory_selection()
            
            # Adım 2: Dosyaları keşfet
            available_files = self.discover_csv_files(selected_directory)
            if not available_files:
                if selected_directory:
                    search_path = self.base_dir / selected_directory
                    print(f"Dizinde CSV/TXT dosyası bulunamadı: {search_path}")
                else:
                    print(f"Ana dizinde CSV/TXT dosyası bulunamadı: {self.base_dir}")
                return False
            
            # Adım 3: Kullanıcı dosya seçimi
            selected_files = self.get_user_file_selection(available_files)
            if not selected_files:
                print("Dönüştürme için dosya seçilmedi")
                return False
            
            # Adım 4: Format tercihi al
            format_type = self.get_excel_format_preference()

            # Adım 5: Sıralama tercihi al
            sort_by_id = self.get_sort_preference()
            
            # Adım 6: Dosyaları oku ve işle
            print(f"\n{len(selected_files)} dosya işleniyor...")
            files_data = {}
            
            for file_path in selected_files:
                data = self.read_csv_file(file_path, sort_by_id)
                if data:
                    files_data[file_path] = data
                    print(f"✓ {file_path.name} ({len(data)} satır)")
                else:
                    print(f"✗ {file_path.name} okunamadı")
            
            # Hiç dosya işlenemediyse dur
            if not files_data:
                print("Hiçbir dosya işlenemedi")
                return False
            
            # Adım 7: Excel çalışma kitabı oluştur
            print("\nExcel çalışma kitabı oluşturuluyor...")
            workbook = self.create_excel_workbook(files_data, format_type)
            if not workbook:
                return False
            
            # Adım 8: Excel dosyasını kaydet
            excel_file = self.save_excel_file(workbook)
            if not excel_file:
                return False
            
            # Adım 9: Arşiv oluştur
            print("Arşiv oluşturuluyor...")
            archive_file = self.create_archive(selected_files, excel_file)
            
            # Adım 10: Sonuçları göster
            self._display_completion_summary(selected_files, excel_file, archive_file)
            
            return True
            
        except KeyboardInterrupt:
            print("\nİşlem kullanıcı tarafından durduruldu")
            self.logger.info("İşlem kullanıcı tarafından durduruldu")
            return False
            
        except Exception as e:
            print(f"\nBeklenmeyen hata: {e}")
            self.logger.error(f"Beklenmeyen hata: {e}")
            return False
    
    def _display_completion_summary(self, source_files: List[Path], 
                                  excel_file: Optional[Path], 
                                  archive_file: Optional[Path]) -> None:
        """Tamamlama özetini kullanıcıya göster"""
        print(f"\n{'='*60}")
        print("DÖNÜŞTÜRME BAŞARIYLA TAMAMLANDI")
        print(f"{'='*60}")
        print(f"İşlenen dosya sayısı: {len(source_files)}")
        
        # Excel dosyası bilgisi
        if excel_file:
            file_size = self._format_file_size(excel_file.stat().st_size)
            print(f"Excel dosyası: {excel_file.name} ({file_size})")
        
        # Arşiv dosyası bilgisi
        if archive_file:
            file_size = self._format_file_size(archive_file.stat().st_size)
            print(f"Arşiv dosyası: {archive_file.name} ({file_size})")
        
        # Konum bilgileri
        print(f"Çıktı dizini: {self.output_dir.absolute()}")
        print(f"Log dosyası: conversion_{self.timestamp}.log")
        print(f"{'='*60}")


def main() -> int:
    """
    Uygulamanın ana giriş noktası
    
    Returns:
        int: Çıkış kodu (başarı için 0, hata için 1)
    """
    try:
        # Dönüştürücü nesnesini oluştur ve çalıştır
        converter = CSVToExcelConverter()
        success = converter.run_conversion()
        return 0 if success else 1
        
    except Exception as e:
        print(f"Kritik hata: {e}")
        return 1


# Script doğrudan çalıştırıldığında main fonksiyonunu çağır
if __name__ == "__main__":
    sys.exit(main())